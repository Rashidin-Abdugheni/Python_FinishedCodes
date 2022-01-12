#
file = open('C:\\Users\\Mr.R\\Desktop\\final\\1.txt', 'r')
text = file.read().strip()
print(text)
file.close()

import pandas as pd
import os
excelName = text
path="C:\\Users\\Mr.R\\Desktop\\test910\\"+excelName
#path=str.replace(path, new[, max])
import sys #print with color
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
prRed(path)
df= pd.read_excel(path)
df.head()
print(df)
#将化学物质的列提到第一列
Name = df['Name']
df.drop(labels=['Name'], axis=1,inplace = True)
df.insert(0, 'Name', Name)
print(df)
newpath="C:\\Users\\Mr.R\\Desktop\\final\\draft\\"+excelName+"final.xlsx"
df.to_excel(newpath)

#将菌株编号替换到peakarea的列名
from openpyxl import load_workbook
path1 = r"C:\\Users\\Mr.R\\Desktop\\final\\draft\\"+excelName+"final.xlsx"
e= load_workbook(path1) #打开excel
E= e.active   #表示当前活跃的表，本案例中 当前活跃表为sheet1
#也可以使用 E = e.get_sheet_by_name('Sheet1') 来获取工作表1
#将excel中 2行3列 对应的数据传给a
a = str(E.cell(row=2, column=4).value)
#b=E.max_row         #读取excel行数
#c=E.max_column     #读取excel列数
print(a)# 将上面获取值给单元格 1,7：
wb=load_workbook(path1) #打开指定excel表
sheet=wb["Sheet1"] #打开指定的Excel表中的sheet页
sheet.cell(1, 9).value=a  #修改第二行第三列的数据为hhh
print(sheet.cell(1,9).value)   #同时工作台打印出被修改的单元格值
update=sheet.cell(1,9).value  #将修改后的值赋值update
print(update)  #update值#### 进行合并 # 1 source table
path2 = r"C:\\Users\\Mr.R\\Desktop\\final\\draft\\"+excelName+"final2.xlsx"
wb.save(path2)

import pandas as pd
dffnl = pd.read_excel("C:\\Users\\Mr.R\\Desktop\\final\\draft\\"+excelName+"final2.xlsx")
print("Here is the df:\n",dffnl)

#Trim the free spaces in front of the data
cols = dffnl.select_dtypes(['object']).columns
dffnl[cols] = dffnl[cols].apply(lambda x: x.str.strip())
#df1=df.sort_index(axis=1)
#df.sort_values(by='ID')
dffnl1 = dffnl.sort_values(by=['Name'])
("Here is the df1:\n",dffnl1)
dffnl1.to_excel("C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx")

# import pandas
import pandas as pd
from openpyxl import *
# read csv data
from openpyxl import load_workbook
path = r"C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx"
e= load_workbook(path) #打开excel
E= e.active   #表示当前活跃的表，本案例中 当前活跃表为sheet1
                    #也可以使用 E = e.get_sheet_by_name('Sheet1') 来获取工作表1
b = str(E.cell(row=2, column=7).value) # 将excel中 2行13列 对应的数据传给a
#b=E.max_row         #读取excel行数
#c=E.max_column     #读取excel列数
print(b)# 将上面获取值给单元格 1,7：
wb=load_workbook("C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx") #打开指定excel表
sheet=wb["Sheet1"] #打开指定的Excel表中的sheet页
sheet.cell(1, 13).value=b  #修改第二行第三列的数据为hhh
wb.save("C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx")  #保存指定Excel工作表
print(sheet.cell(1, 13).value)   #同时工作台打印出被修改的单元格值
update=sheet.cell(1, 13).value  #将修改后的值赋值update
print(update)  #update值#### 进行合并 # 1 source table
df1 = pd.read_excel("C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx")
print(df1)
#上面已经得到了chem在第一列的没问题的数据！
#vlookup
import pandas as pd
# 地区信息表
strain=excelName
df_location = pd.read_excel("D:\\20210911-FINAL\\Vlookup-SCFA-SPME-list-part1-913.xlsx" )
df_location.head()
# 数据库导出表
strain_path="C:\\Users\\Mr.R\\Desktop\\final\\ChemFirst\\"+excelName+"-ChemFirst.xlsx"
df_number = pd.read_excel(strain_path)
df_number.head()
print(df_number)
# 只筛选第二个表的少量的列（只选取表二中市区和用户人数）
df_number = df_number[[ "Name", b]]
df_number.head()
df_merge = pd.merge(left=df_location, right=df_number, left_on="Name", right_on="Name",how='outer')
df_merge.head()
df_merge.to_excel("D:\\20210911-FINAL\\Vlookup-SCFA-SPME-list-part1-913.xlsx", index=False)