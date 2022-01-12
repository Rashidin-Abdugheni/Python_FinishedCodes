
from openpyxl import load_workbook

wb = load_workbook('D:\\USB-GCMS-FILES\\SCFA-GCMS-EXCEL-RESULT_OF_EACH_STRAIN\\SD-N-D8-16-SCFAs-结果.xlsx')
ws = wb['Sheet1']
ws.cell(row=1,column=1).value = 'Strain'
ws.cell(row=1,column=2).value = 'No'
ws.cell(row=1,column=3).value = 'RetentionTime'
ws.cell(row=1,column=4).value = 'Starting'
ws.cell(row=1,column=5).value = 'Ending'
ws.cell(row=1,column=6).value = 'm/z'
ws.cell(row=1,column=7).value = 'PeakArea'

ws.cell(row=1,column=8).value = 'Peak%'
ws.cell(row=1,column=9).value = 'PeakHight'
ws.cell(row=1,column=10).value = 'PeakHight%'
ws.cell(row=1,column=11).value = 'A/H'
ws.cell(row=1,column=12).value = 'Node'
ws.cell(row=1,column=13).value = 'Name'


wb.save('D:\\USB-GCMS-FILES\\SCFA-GCMS-EXCEL-RESULT_OF_EACH_STRAIN\\SD-N-D8-16-SCFAs-English.xlsx')
