# 重要：安装xlrd时要安装以下这个！
# pip install xlrd==1.2.0
# pip install pandas

import os
import pandas as pd
import numpy as np


dir= "D:\\USB-GCMS-FILES\\MergeTrims"#设置工作路径，读取其中的多个excel表

#新建列表，存放文件名（可以忽略，但是为了做的过程能心里有数，先放上）
filename_excel = []

#新建列表，存放每个文件数据框（每一个excel读取后存放在数据框）
frames = []

for root, dirs, files in os.walk(dir):
    for file in files:

        filename_excel.append(os.path.join(root,file))

        # excel转换成DataFrame
        df = pd.read_excel(os.path.join(root,file))
        frames.append(df)
#打印文件名
print(filename_excel)

 #合并所有数据
result = pd.concat(frames)

#查看合并后的数据
result.head()
result.shape

#合并文件输出
result.to_csv('D:\\USB-GCMS-FILES\\Merge-all.csv', sep=',', index=False)

#3 Transfer csv into xlsx

import openpyxl
import csv

wb = openpyxl.Workbook()
sh = wb.create_sheet('GCMS-DATA')

with open('D:\\USB-GCMS-FILES\\Merge-all.csv', 'r') as f:
    reader = csv.reader(f)
    for row in reader:
        sh.append(row)

wb.save('D:\\USB-GCMS-FILES\\Merge-English-all.xlsx')



import pandas as pd

df = pd.read_excel('D:\\USB-GCMS-FILES\\Merge-English-all.xlsx',sheet_name='GCMS-DATA')

cols = df.select_dtypes(['object']).columns # Trim the free spaces in front of the data
df[cols] = df[cols].apply(lambda x: x.str.strip())
print (df)


#df1=df.sort_index(axis=1)
#print(df1)

#df.sort_values(by='ID')
#print(",\,")
#df1=df.sort_values(by=['Name'])
#print(df1)

df.to_excel("D:\\USB-GCMS-FILES\\Merge-English-all-Final.xlsx")


#from openpyxl import *

#filename = r'D:\\USB-GCMS-FILES\\SCFA-SPME-OF_EACH_STRAIN-English-all-Sorted.xlsx'
#wb = load_workbook(filename)
#ws = wb.active
#ws.delete_cols(15) #删除第 15 列数据
#ws.delete_cols(16)
#ws.delete_cols(17)
#ws.delete_cols(18)

#wb.save(filename)