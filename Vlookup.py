import pandas as pd # import pandas
from openpyxl import * #读取某个单元格的值
from openpyxl import load_workbook
path = r'G:/GCMS-TABLE/Splited_SPME0307/Sheet48-.xlsx'
e= load_workbook(path) #打开excel
E= e.active   #表示当前活跃的表，本案例中 当前活跃表为sheet1  #也可以使用 E = e.get_sheet_by_name('Sheet1') 来获取工作表1
a = str(E.cell(row=2, column=13).value) # 将excel中 1行2列 对应的数据传给a#b=E.max_row   #读取excel行数 #c=E.max_column     #读取excel列数
print(a)
wb=load_workbook("G:/GCMS-TABLE/Splited_SPME0307/Sheet48-.xlsx") #打开指定excel表
sheet=wb["Sheet48"] #打开指定的Excel表中的sheet页
sheet.cell(1,7).value=a  #修改第二行第三列的数据为hhh
wb.save("G:/GCMS-TABLE/Splited_SPME0307/Sheet48-.xlsx")  #保存指定Excel工作表
print(sheet.cell(2,3).value)   #同时工作台打印出被修改的单元格值
update=sheet.cell(2,3).value  #将修改后的值赋值update
print(update)  #update值
filename = r'G:/GCMS-TABLE/Splited_SPME0307/Sheet48-.xlsx'
wb = load_workbook(filename)
ws = wb.active
ws.delete_cols(2,5) #删除第 13 列数据 #ws.delete_rows(3) #删除第 3行数据
ws.delete_cols(3,10) #删除列数据
wb.save(filename)# 写入到制定单元格
df1 = pd.read_excel('G:/GCMS-TABLE/Splited_SPME0307/Sheet48-.xlsx')#source table
print(df1)
df2 = pd.read_excel('G:/GCMS-TABLE/Vlookuped/SPME-MergedFile1.xlsx')# Blank table
Left_join = pd.merge(df1, df2, on='Unnamed: 0', how='outer')
print(Left_join)
df = pd.DataFrame(Left_join)
gfg_csv_data = df.to_excel('G:/GCMS-TABLE/Vlookuped/SPME-MergedFile1.xlsx', index =False )

