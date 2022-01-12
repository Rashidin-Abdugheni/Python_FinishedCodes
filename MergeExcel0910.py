# 重要：安装xlrd时要安装以下这个！
# pip install xlrd==1.2.0
# pip install pandas
import os
import pandas as pd
import numpy as np
dir= "D:\\20210911-FINAL\\scfa"#设置工作路径，读取其中的多个excel表
#新建列表，存放文件名（可以忽略，但是为了做的过程能心里有数，先放上）
filename_excel = []
#新建列表，存放每个文件数据框（每一个excel读取后存放在数据框）
frames = []
for root, dirs, files in os.walk(dir):
    for file in files:
        filename_excel.append(os.path.join(root,file))
        # excel转换成DataFrame
        df = pd.read_excel(os.path.join(root,file))
        frames.append(df)
#打印文件名
print("1 Here is the Merged filename_excel:", filename_excel)
 #合并所有数据
result = pd.concat(frames)
#查看合并后的数据
result.head()
result.shape
#合并文件输出
pathdir='D:\\20210911-FINAL\\910-1-Merged.csv'
result.to_csv(pathdir, sep=',', index=False)

#转换为xlsx
import openpyxl
import csv
wb = openpyxl.Workbook()
sh = wb.create_sheet('Sheet1')
with open(pathdir, 'r') as f:
    reader = csv.reader(f)
    for row in reader:
        sh.append(row)
        transferred_path1='D:\\20210911-FINAL\\910-1-Merged.xlsx'
wb.save(transferred_path1)
#根据列值排序
import pandas as pd
df = pd.read_excel(transferred_path1,sheet_name='Sheet1')
print("2 Here is transferred Excel:",df)
#df1=df.sort_index(axis=1)
#print(df1)
#df.sort_values(by='ID')
print(",\,")
df1=df.sort_values(by=['Name'])
print("Sorted Merge:",df1)
sorted_path='D:\\20210911-FINAL\\910-3-Sorted.xlsx'
df1.to_excel(sorted_path)

#提取指定的列名称的为新的文件夹
import pandas as pd
path = pathdir
# 使用pandas读入
data = pd.read_excel(sorted_path,sheet_name='Sheet1') #读取文件中所有数据
# 按列分离数据
x = data[['Name']]#读取某1列
Name_path='D:\\20210911-FINAL\\910-4_NameColumn.csv'
x.to_csv(Name_path)
#print(# 第Name列
print("Name Column only table content:",x)
#name转换为xlsx
import openpyxl
import csv
wb = openpyxl.Workbook()
sh = wb.create_sheet('Sheet1')
with open(Name_path, 'r') as f:
    reader = csv.reader(f)
    for row in reader:
        sh.append(row)
transferred_path2='D:\\20210911-FINAL\\910-4_NameColumn.xlsx'
wb.save(transferred_path2)
print("Transferred Excel of Name Column")
#删除重复值
import pandas as pd
Dropduplicates = pd.read_excel(transferred_path2,sheet_name='Sheet1')
Dropduplicates.drop_duplicates(subset=["Name"], keep="first")
print(Dropduplicates)
path_Dropduplicates='D:\\20210911-FINAL\\Dropduplicates.xlsx'
Dropduplicates.to_excel(path_Dropduplicates)

#输出带颜色的字段
import sys
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
def prGreen(skk): print("\033[92m {}\033[00m" .format(skk))
def prYellow(skk): print("\033[93m {}\033[00m" .format(skk))
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk))
def prPurple(skk): print("\033[95m {}\033[00m" .format(skk))
def prCyan(skk): print("\033[96m {}\033[00m" .format(skk))
def prLightGray(skk): print("\033[97m {}\033[00m" .format(skk))
def prBlack(skk): print("\033[98m {}\033[00m" .format(skk))
prGreen("\n DropDuolicates is done!")



