import pandas as pd

# 地区信息表
strain="SPME-20210809-10-T296-44"
df_location = pd.read_excel("C:\\Users\\Mr.R\\Desktop\\final\\All_chemicals_Dropduplicated.xlsx" )
df_location.head()

# 数据库导出表
strain_path="D:\\20210911-FINAL\merged\Taxon-dropduplicated\\"+strain+".xlsx"
df_number = pd.read_excel(strain_path)
df_number.head()
print(df_number)
from openpyxl import load_workbook
path1 = r"D:\\20210911-FINAL\\merged\\Taxon-dropduplicated\\"+strain+".xlsx"
e= load_workbook(path1) #打开excel
E= e.active   #表示当前活跃的表，本案例中 当前活跃表为sheet1
#也可以使用 E = e.get_sheet_by_name('Sheet1') 来获取工作表1
#将excel中 2行3列 对应的数据传给a
a = str(E.cell(row=2, column=4).value)
#b=E.max_row         #读取excel行数
#c=E.max_column     #读取excel列数
print("a is :", a)# 将上面获取值给单元格 1,7：

# 只筛选第二个表的少量的列（只选取表二中市区和用户人数）
df_number = df_number[[ "Name", a]]
df_number.head()

df_merge = pd.merge(left=df_location, right=df_number, left_on="Name", right_on="Name",how='outer')
df_merge.head()

df_merge.to_excel("C:\\Users\\Mr.R\\Desktop\\final\\All_chemicals_Dropduplicated.xlsx", index=False)
